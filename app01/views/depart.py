from django.shortcuts import render, redirect, HttpResponse
from app01 import models

from app01.utils.pagination import Pagination
from app01.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm


def depart_list(request):
    """ Department List """

    # Retrieve all department entries from the database
    #  [Object, Object, Object]
    queryset = models.Department.objects.all()

    return render(request, 'depart_list.html', {'queryset': queryset})


def depart_add(request):
    """ Add Department """
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # Get data submitted via POST (title is empty)
    title = request.POST.get("title")

    # Save to database
    models.Department.objects.create(title=title)

    # Redirect back to department list
    return redirect("/depart/list/")


def depart_delete(request):
    """ Delete Department """
    # Get ID from URL http://127.0.0.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')

    # Delete entry
    models.Department.objects.filter(id=nid).delete()

    # Redirect back to department list
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """ Edit Department """
    if request.method == "GET":
        # Retrieve data by nid [obj,]
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {"row_object": row_object})

    # Get title submitted by user
    title = request.POST.get("title")

    # Find entry in the database by ID and update
    # models.Department.objects.filter(id=nid).update(title=title,other_field=123)
    models.Department.objects.filter(id=nid).update(title=title)

    # Redirect back to department list
    return redirect("/depart/list/")


def depart_multi(request):
    """ Bulk Add Departments (Excel File) """
    from openpyxl import load_workbook

    # 1. Get file object uploaded by user
    file_object = request.FILES.get("exc")

    # 2. Pass the object to openpyxl, let openpyxl read file content
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3. Loop to get each row of data
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')

